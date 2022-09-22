# -*- coding: utf-8 -*-
'''
@Time :  16:04
@Author : maodongdong
@File : ExcelDate.py
'''
import json
import xlrd as xlrd
from openpyxl import load_workbook

from Common.getData import GetData
from project_path import project_path, test_case_path, case_config_path
from read_config import ReadConfig

'''
读取excel数据
'''

class Handle_Excel:
    # def __init__(self):
    #     self.excelDir = r'D:\test_automation\test_automation\TestDatas\Login\testcase.xls'

    # def get_excel_data(self,sheetName,caseName):
    #     '''
    #     :param sheetName: 表名
    #     :param caseName: 用例名
    #     :return: 一个列表嵌入元组    [(请求体1，期望数据1),(请求体2，期望数据2)]
    #     '''
    #     resultList = []  #放结果的
    #
    #     # excelDir = '..../TestDatas/Login/testcase.xls'
    #     # 1：excel 对象 保持格式
    #     workBook = xlrd.open_workbook(self.excelDir,formatting_info=True)
    #     # 2：针对某个sheet操作
    #     workSheet = workBook.sheet_by_name(sheetName)
    #     # 3：获取值 第4列和第5列
    #     # print(workSheet.col_values(0))
    #     #4:获取数据
    #     idx = 0
    #     for one in workSheet.col_values(0):
    #         if caseName in one:
    #             #请求体-行号和列号重新开始
    #             resWay = workSheet.cell(idx,4).value
    #             reqBody = workSheet.cell(idx,6).value
    #             respData = workSheet.cell(idx,8).value #期望数据
    #             #j将每行数据返回至List中
    #             #字符串转换成字典   json.loads(json数据)
    #             resultList.append((resWay,json.loads(reqBody),json.loads(respData)))#字符串格式
    #         idx += 1
    #     return resultList
    @classmethod
    def get_excel_data(self,fileName):
        '''
        :param sheetName: 表名
        :param caseName: 用例名
        :return: 一个列表嵌入元组    [(请求体1，期望数据1),(请求体2，期望数据2)]
        '''
        workBook = load_workbook(fileName)
        mode = eval(ReadConfig.get_config(case_config_path, 'MODE', 'mode'))
        resultList = []  #放结果的

        # workSheet = workBook[sheetName]
        tel = GetData.NoRegTel
        for key in mode:
            workSheet = workBook[key]
            if mode[key] == "all":
                for i in range(2, workSheet.max_row + 1):
                    row_data = {}
                    row_data['case_id'] = workSheet.cell(i, 1).value
                    row_data['case_name'] = workSheet.cell(i, 2).value
                    row_data['case_des'] = workSheet.cell(i, 3).value
                    row_data['case_url'] = workSheet.cell(i, 4).value
                    row_data['case_way'] = workSheet.cell(i, 5).value
                    row_data['case_handle'] = workSheet.cell(i, 6).value
                    # row_data['case_data'] = workSheet.cell(i, 7).value
                    if workSheet.cell(i,7).value.find('${tel_1}') != -1:
                        row_data['case_data'] = workSheet.cell(i, 7).value.replace('${tel_1}',str(tel))
                    elif workSheet.cell(i,7).value.find('${tel}') != -1:
                        row_data['case_data'] = workSheet.cell(i, 7).value.replace('${tel}', str(tel+1))
                    else:
                        row_data['case_data'] = workSheet.cell(i, 7).value
                    row_data['case_expectData'] = workSheet.cell(i, 8).value
                    row_data['sheet_name'] = key
                    resultList.append(row_data)
                    self.update_excel(fileName,'init',tel+2)
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = workSheet.cell(float(case_id) + 1, 1).value
                    row_data['case_name'] = workSheet.cell(case_id + 1, 2).value
                    row_data['case_des'] = workSheet.cell(case_id + 1, 3).value
                    row_data['case_url'] = workSheet.cell(case_id + 1, 4).value
                    row_data['case_way'] = workSheet.cell(case_id + 1, 5).value
                    row_data['case_handle'] = workSheet.cell(case_id + 1, 6).value
                    # row_data['case_data'] = workSheet.cell(case_id + 1, 7).value
                    if workSheet.cell(case_id + 1,7).value.find('${tel_1}') != -1:
                        row_data['case_data'] = workSheet.cell(case_id + 1, 7).value.replace('${tel_1}',str(tel))
                    elif workSheet.cell(case_id + 1,7).value.find('${tel}') != -1:
                        row_data['case_data'] = workSheet.cell(case_id + 1, 7).value.replace('${tel}', str(tel+1))
                    else:
                        row_data['case_data'] = workSheet.cell(case_id + 1, 7).value
                    row_data['case_expectData'] = workSheet.cell(case_id+1, 8).value
                    row_data['sheet_name'] = key
                    resultList.append(row_data)
                    self.update_excel(fileName, 'init', tel + 2)

        return resultList

    @staticmethod
    def do_excel_writeback(fileName,sheetName,i,actual_data,result):
        workBook = load_workbook(fileName)
        sheet = workBook[sheetName]
        sheet.cell(i,10).value = actual_data
        sheet.cell(i,11).value = result
        workBook.save(fileName)

    @classmethod
    def update_excel(self,fileName,sheetName,tel):
        workBook = load_workbook(fileName)
        sheet = workBook[sheetName]
        sheet.cell(2,1).value = tel
        workBook.save(fileName)


if __name__ =='__main__':
    test_data = Handle_Excel().get_excel_data(test_case_path)
    print(test_data)






# class RunTestCase:
#     def __init__(self):
#         self.Runmain = RunMain()  # 实例化调用get/post请求基类
#         self.data = HandleExcel()  # 实例化操作excel文件类
#
#     def go_run(self):
#         rows_count = self.data.get_rows()   # 获取excel行数
#         for i in range(1,rows_count):      # 利用行数进行迭代处理每个接口
#             url = self.data.get_value(i, get_url())  # 循环获取url的值
#             # print(url)
#             method = self.data.get_value(i, get_method())  # 循环获取method的值
#             data = json.loads(self.data.get_value(i, get_params()))   # 循环获取请求参数，并将得到的数据反序列化
#             # data = self.data.get_value(i, get_params())  # 循环获取请求参数
#             print(data)
#             is_run = self.data.get_value(i, get_priority())  # 获取是否运行，即判断excel中priority是不是"H"
#             if is_run == 'H':
#                 res = self.Runmain.run_main(url, method, data)  # 调用get/post主函数
#                 print(res)
#
